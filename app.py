from flask_caching import Cache
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
from flask_mail import Mail, Message
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from database import init_db
from datetime import datetime
import os

DB_CONFIG = {
    'host': os.environ.get('DB_HOST', 'aws-1-us-east-2.pooler.supabase.com'),
    'port': int(os.environ.get('DB_PORT', 5432)),
    'database': os.environ.get('DB_NAME', 'postgres'),
    'user': os.environ.get('DB_USER', 'postgres.alvxqmzaiocvmdkjgpqh'),
    'password': os.environ.get('DB_PASSWORD', '')
}

import psycopg2
import psycopg2.extras
from psycopg2 import pool as pg_pool

# Pool de conexiones global
connection_pool = None

def init_pool():
    global connection_pool
    connection_pool = pg_pool.ThreadedConnectionPool(
        minconn=2,
        maxconn=10,
        **DB_CONFIG
    )

def get_db():
    global connection_pool
    if connection_pool is None:
        init_pool()
    
    conn = connection_pool.getconn()

    class RowWrapper:
        def __init__(self, row):
            self._row = dict(row) if row else {}
            self._keys = list(self._row.keys())
        def __getitem__(self, key):
            if isinstance(key, int):
                return self._row[self._keys[key]]
            return self._row[key]
        def __iter__(self):
            return iter(self._row.values())
        def keys(self):
            return self._keys
        def get(self, key, default=None):
            return self._row.get(key, default)
        def __contains__(self, key):
            return key in self._row
        def items(self):
            return self._row.items()
        def values(self):
            return self._row.values()

    class CursorWrapper:
        def __init__(self, cursor):
            self._cursor = cursor
        def fetchone(self):
            row = self._cursor.fetchone()
            return RowWrapper(row) if row else None
        def fetchall(self):
            return [RowWrapper(r) for r in self._cursor.fetchall()]
        def __iter__(self):
            for row in self._cursor:
                yield RowWrapper(row)

    class DBWrapper:
        def __init__(self, connection):
            self.conn = connection
            self._cursor = connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        def execute(self, query, params=None):
            # Crear nuevo cursor para cada consulta
            cursor = self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            query = query.replace('?', '%s')
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            return CursorWrapper(cursor)
        def executescript(self, script):
            self._cursor.execute(script)
        def commit(self):
            self.conn.commit()
        def close(self):
            self._cursor.close()
            connection_pool.putconn(self.conn)

    return DBWrapper(conn)
app = Flask(__name__)
# Configuración de caché
app.config['CACHE_TYPE'] = 'SimpleCache'
app.config['CACHE_DEFAULT_TIMEOUT'] = 300  # 5 minutos
cache = Cache(app)
app.secret_key = os.environ.get('SECRET_KEY', 'medline-almacen-2024-clave-secreta')

# ─── CONFIGURACIÓN DE CORREO ──────────────────────────────────────────────────
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.secret_key = os.environ.get('SECRET_KEY', 'medline-almacen-2024-clave-secreta')
app.config['MAIL_USERNAME'] = os.environ.get('almacen.medline.alertas@gmail.com', '')
app.config['MAIL_PASSWORD'] = os.environ.get('dqhubredptxbabur', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('almacen.medline.alertas@gmail.com', '')
mail = Mail(app)

@app.context_processor
def inject_now():
    alertas_count = 0
    if 'usuario_id' in session:
        try:
            db = get_db()
            rol = session.get('rol')
            id_proveedor = session.get('id_proveedor')
            if rol == 'admin':
                alertas_count = db.execute(
                    'SELECT COUNT(*) FROM alertas WHERE resuelta=0'
                ).fetchone()[0]
            else:
                alertas_count = db.execute('''
                    SELECT COUNT(*) FROM alertas a
                    JOIN articulos art ON a.id_articulo = art.id_articulo
                    WHERE a.resuelta=0 AND art.id_proveedor=?
                ''', (id_proveedor,)).fetchone()[0]
            db.close()
        except:
            pass
    from datetime import datetime
    return {'now': datetime.now(), 'now_str': datetime.now().strftime('%m/%d/%Y %I:%M %p'),
            'alertas_count': alertas_count}

UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_requerido(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def solo_admin(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('rol') != 'admin':
            flash('Acceso restringido al administrador.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

# ─── RUTAS DE AUTENTICACIÓN ───────────────────────────────────────────────────

@app.route('/')
def index():
    if 'usuario_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db = get_db()
        usuario = db.execute(
            'SELECT * FROM usuarios WHERE username = ? AND activo = 1', (username,)
        ).fetchone()
        db.close()
        if usuario and check_password_hash(usuario['password_hash'], password):
            session['usuario_id'] = usuario['id_usuario']
            session['nombre'] = usuario['nombre']
            session['rol'] = usuario['rol']
            session['id_proveedor'] = usuario['id_proveedor']
            return redirect(url_for('dashboard'))
        flash('Usuario o contraseña incorrectos.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/kiosco')
def kiosco():
    return render_template('kiosco.html')

# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_requerido
def dashboard():
    db = get_db()
    rol = session.get('rol')
    id_proveedor = session.get('id_proveedor')

    if rol == 'admin':
        total_articulos = db.execute('SELECT COUNT(*) FROM articulos WHERE activo=1').fetchone()[0]
        total_proveedores = db.execute('SELECT COUNT(*) FROM proveedores WHERE activo=1').fetchone()[0]
        alertas = db.execute('''
            SELECT a.*, art.nombre as nombre_articulo, art.stock_actual, art.stock_minimo
            FROM alertas a
            JOIN articulos art ON a.id_articulo = art.id_articulo
            WHERE a.resuelta = 0 ORDER BY a.fecha_hora DESC LIMIT 10
        ''').fetchall()
        criticos = db.execute(
            'SELECT * FROM articulos WHERE es_critico=1 AND activo=1'
        ).fetchall()
        movimientos_recientes = db.execute('''
            SELECT m.*, art.nombre as nombre_articulo, u.nombre as nombre_usuario
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            LEFT JOIN usuarios u ON m.id_usuario = u.id_usuario
            ORDER BY m.fecha_hora DESC LIMIT 10
        ''').fetchall()
    else:
        total_articulos = db.execute(
            'SELECT COUNT(*) FROM articulos WHERE activo=1 AND id_proveedor=?', (id_proveedor,)
        ).fetchone()[0]
        total_proveedores = None
        alertas = db.execute('''
            SELECT a.*, art.nombre as nombre_articulo, art.stock_actual, art.stock_minimo
            FROM alertas a
            JOIN articulos art ON a.id_articulo = art.id_articulo
            WHERE a.resuelta = 0 AND art.id_proveedor = ?
            ORDER BY a.fecha_hora DESC LIMIT 10
        ''', (id_proveedor,)).fetchall()
        criticos = db.execute(
            'SELECT * FROM articulos WHERE es_critico=1 AND activo=1 AND id_proveedor=?',
            (id_proveedor,)
        ).fetchall()
        movimientos_recientes = db.execute('''
            SELECT m.*, art.nombre as nombre_articulo, u.nombre as nombre_usuario
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            LEFT JOIN usuarios u ON m.id_usuario = u.id_usuario
            WHERE art.id_proveedor = ?
            ORDER BY m.fecha_hora DESC LIMIT 10
        ''', (id_proveedor,)).fetchall()

    db.close()
    # Convertir movimientos a lista serializable para JSON
    movimientos_json = []
    for m in movimientos_recientes:
        movimientos_json.append({
            'tipo': m['tipo'],
            'cantidad': m['cantidad'],
            'fecha_hora': str(m['fecha_hora']) if m['fecha_hora'] else ''
        })

    # Top articulos mas consumidos para grafica
    if rol == 'admin':
        top_consumo = db.execute('''
            SELECT art.nombre, 
                   SUM(CASE WHEN m.tipo='SALIDA' THEN m.cantidad ELSE 0 END) as total_salidas,
                   SUM(CASE WHEN m.tipo='ENTRADA' THEN m.cantidad ELSE 0 END) as total_entradas
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            GROUP BY art.id_articulo, art.nombre
            ORDER BY total_salidas DESC LIMIT 8
        ''').fetchall()
    else:
        top_consumo = db.execute('''
            SELECT art.nombre,
                   SUM(CASE WHEN m.tipo='SALIDA' THEN m.cantidad ELSE 0 END) as total_salidas,
                   SUM(CASE WHEN m.tipo='ENTRADA' THEN m.cantidad ELSE 0 END) as total_entradas
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            WHERE art.id_proveedor = %s
            GROUP BY art.id_articulo, art.nombre
            ORDER BY total_salidas DESC LIMIT 8
        ''', (id_proveedor,)).fetchall()

    top_consumo_json = [{'nombre': r['nombre'], 'salidas': r['total_salidas'], 'entradas': r['total_entradas']} for r in top_consumo]
    
    return render_template('dashboard.html',
        total_articulos=total_articulos,
        total_proveedores=total_proveedores,
        alertas=alertas,
        criticos=criticos,
        movimientos_recientes=movimientos_recientes,
        top_consumo_json=top_consumo_json,
        movimientos_json=movimientos_json
    )

# ─── ARTÍCULOS ────────────────────────────────────────────────────────────────

@app.route('/articulos')
@login_requerido
@cache.cached(timeout=60, key_prefix=lambda: f'articulos_{session.get("usuario_id")}')
def articulos():
    db = get_db()
    rol = session.get('rol')
    id_proveedor = session.get('id_proveedor')
    if rol == 'admin':
        lista = db.execute('''
            SELECT a.*, p.nombre as nombre_proveedor
            FROM articulos a JOIN proveedores p ON a.id_proveedor = p.id_proveedor
            WHERE a.activo = 1 ORDER BY a.nombre
        ''').fetchall()
    else:
        lista = db.execute('''
            SELECT a.*, p.nombre as nombre_proveedor
            FROM articulos a JOIN proveedores p ON a.id_proveedor = p.id_proveedor
            WHERE a.activo = 1 AND a.id_proveedor = ? ORDER BY a.nombre
        ''', (id_proveedor,)).fetchall()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()
    db.close()
    return render_template('articulos.html', articulos=lista, proveedores=proveedores)

@app.route('/articulos/nuevo', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def nuevo_articulo():
    db = get_db()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()
    if request.method == 'POST':
        cache.clear()
        imagen_filename = None
        if 'imagen' in request.files:
            file = request.files['imagen']
            if file and allowed_file(file.filename):
                imagen_filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], imagen_filename))
        num_parte = request.form['num_parte']
        db.execute('''
            INSERT INTO articulos (num_parte, nombre, descripcion, unidad_medida,
                stock_minimo, codigo_qr, link_compra, imagen, id_proveedor, ubicacion)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            num_parte,
            request.form['nombre'],
            request.form['descripcion'],
            request.form['unidad_medida'],
            request.form['stock_minimo'],
            num_parte,
            request.form['link_compra'],
            imagen_filename,
            request.form['id_proveedor'],
            request.form.get('ubicacion') or None
        ))
        db.commit()
        db.close()
        flash('Artículo creado correctamente.', 'success')
        return redirect(url_for('articulos'))
    db.close()
    return render_template('nuevo_articulo.html', proveedores=proveedores)

# ─── MOVIMIENTOS / ESCANEO ────────────────────────────────────────────────────

@app.route('/registrar_movimiento', methods=['POST'])
def registrar_movimiento():
    data = request.get_json()
    codigo_qr = data.get('codigo_qr')
    tipo = data.get('tipo')
    cantidad = int(data.get('cantidad', 1))
    observaciones = data.get('observaciones', '')
    modo_kiosco = data.get('modo_kiosco', 0)
    id_usuario = session.get('usuario_id')

    db = get_db()
    articulo = db.execute(
        'SELECT * FROM articulos WHERE codigo_qr = ? AND activo = 1', (codigo_qr,)
    ).fetchone()

    if not articulo:
        db.close()
        return jsonify({'status': 'error', 'message': 'Artículo no encontrado.'})

    # Solo admin puede hacer entradas
    rol = session.get('rol', 'kiosco')
    if tipo == 'ENTRADA' and rol not in ['admin']:
        db.close()
        return jsonify({'status': 'error', 'message': 'Solo el administrador puede registrar entradas.'})

    if tipo == 'SALIDA' and articulo['stock_actual'] < cantidad:
        db.close()
        return jsonify({'status': 'error', 'message': 'Stock insuficiente.'})

    db.execute('''
        INSERT INTO movimientos (id_articulo, id_usuario, tipo, cantidad, observaciones, modo_kiosco)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (articulo['id_articulo'], id_usuario, tipo, cantidad, observaciones, modo_kiosco))

    nuevo_stock = articulo['stock_actual'] + cantidad if tipo == 'ENTRADA' else articulo['stock_actual'] - cantidad
    db.execute('UPDATE articulos SET stock_actual = ? WHERE id_articulo = ?',
               (nuevo_stock, articulo['id_articulo']))

    if nuevo_stock <= articulo['stock_minimo']:
        db.execute('''
            INSERT INTO alertas (id_articulo, tipo_alerta)
            VALUES (?, ?)
        ''', (articulo['id_articulo'], 'STOCK_MINIMO'))

        try:
            admin = db.execute(
                "SELECT * FROM usuarios WHERE rol='admin' AND activo=1"
            ).fetchone()

            proveedor_usuario = db.execute('''
                SELECT u.* FROM usuarios u
                JOIN articulos art ON u.id_proveedor = art.id_proveedor
                WHERE art.id_articulo = ? AND u.rol = 'proveedor' AND u.activo = 1
                LIMIT 1
            ''', (articulo['id_articulo'],)).fetchone()

            import json
            destinatarios = []
            if admin and admin['email']:
                destinatarios.append(admin['email'])
            if proveedor_usuario and proveedor_usuario['email']:
                if proveedor_usuario['email'] not in destinatarios:
                    destinatarios.append(proveedor_usuario['email'])
            # Correo adicional de configuración
            try:
                if os.path.exists('config.json'):
                    with open('config.json', 'r') as f:
                        cfg = json.load(f)
                    correo_extra = cfg.get('correo_alertas', '')
                    if correo_extra and correo_extra not in destinatarios:
                        destinatarios.append(correo_extra)
            except:
                pass

            if destinatarios:
                msg = Message(
                    subject=f'⚠️ Alerta de stock bajo - {articulo["nombre"]}',
                    recipients=destinatarios,
                    html=f'''
                    <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;">
                      <div style="background:#0b4f6c;padding:20px;border-radius:10px 10px 0 0;">
                        <h2 style="color:#fff;margin:0;">⚠️ Alerta de Stock Bajo</h2>
                        <p style="color:#cce7f0;margin:5px 0 0;">Sistema de Almacén Medline</p>
                      </div>
                      <div style="background:#fff;padding:24px;border:1px solid #eee;border-radius:0 0 10px 10px;">
                        <p>El siguiente artículo ha alcanzado su stock mínimo:</p>
                        <table style="width:100%;border-collapse:collapse;margin:16px 0;">
                          <tr style="background:#f0f8fc;">
                            <td style="padding:10px;font-weight:bold;">Artículo</td>
                            <td style="padding:10px;">{articulo["nombre"]}</td>
                          </tr>
                          <tr>
                            <td style="padding:10px;font-weight:bold;">Número de parte</td>
                            <td style="padding:10px;">{articulo["num_parte"]}</td>
                          </tr>
                          <tr style="background:#f0f8fc;">
                            <td style="padding:10px;font-weight:bold;">Stock actual</td>
                            <td style="padding:10px;color:#dc3545;font-weight:bold;">{nuevo_stock}</td>
                          </tr>
                          <tr>
                            <td style="padding:10px;font-weight:bold;">Stock mínimo</td>
                            <td style="padding:10px;">{articulo["stock_minimo"]}</td>
                          </tr>
                        </table>
                        <p style="color:#888;font-size:13px;">
                          Por favor genera una orden de compra para reabastecer este artículo.
                        </p>
                      </div>
                    </div>
                    '''
                )
                mail.send(msg)
        except Exception as e:
            print(f"Error enviando correo: {e}")
# Resolver alertas si stock subió sobre el mínimo
    if tipo == 'ENTRADA' and nuevo_stock > articulo['stock_minimo']:
        db.execute(
            'UPDATE alertas SET resuelta=1 WHERE id_articulo=%s AND resuelta=0',
            (articulo['id_articulo'],)
        )
    db.commit()
    db.close()
    return jsonify({
        'status': 'success',
        'message': f'Movimiento registrado. Stock actual: {nuevo_stock}',
        'articulo': articulo['nombre'],
        'stock_actual': nuevo_stock
    })

# ─── PROVEEDORES ──────────────────────────────────────────────────────────────

@app.route('/proveedores')
@login_requerido
@solo_admin
@cache.cached(timeout=120, key_prefix='proveedores')
def proveedores():
    db = get_db()
    lista = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()
    db.close()
    return render_template('proveedores.html', proveedores=lista)

@app.route('/proveedores/nuevo', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def nuevo_proveedor():
    if request.method == 'POST':
        db = get_db()
        db.execute('INSERT INTO proveedores (nombre, contacto, area) VALUES (?, ?, ?)',
                   (request.form['nombre'], request.form['contacto'], request.form['area']))
        db.commit()
        db.close()
        flash('Proveedor creado correctamente.', 'success')
        return redirect(url_for('proveedores'))
    return render_template('nuevo_proveedor.html')
@app.route('/proveedores/editar/<int:id_proveedor>', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def editar_proveedor(id_proveedor):
    db = get_db()
    proveedor = db.execute('SELECT * FROM proveedores WHERE id_proveedor=?', (id_proveedor,)).fetchone()
    if request.method == 'POST':
        logo_filename = proveedor['logo'] if 'logo' in proveedor.keys() else None
        if 'logo' in request.files:
            file = request.files['logo']
            if file and allowed_file(file.filename):
                logo_filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], logo_filename))
        db.execute('''
            UPDATE proveedores SET nombre=?, contacto=?, area=?, logo=?
            WHERE id_proveedor=?
        ''', (
            request.form['nombre'],
            request.form['contacto'],
            request.form['area'],
            logo_filename,
            id_proveedor
        ))
        db.commit()
        db.close()
        flash('Proveedor actualizado correctamente.', 'success')
        return redirect(url_for('proveedores'))
    db.close()
    return render_template('editar_proveedor.html', proveedor=proveedor)

@app.route('/proveedores/desactivar/<int:id_proveedor>')
@login_requerido
@solo_admin
def desactivar_proveedor(id_proveedor):
    db = get_db()
    db.execute('UPDATE proveedores SET activo=0 WHERE id_proveedor=?', (id_proveedor,))
    db.commit()
    db.close()
    flash('Proveedor desactivado.', 'success')
    return redirect(url_for('proveedores'))

@app.route('/articulos/editar/<int:id_articulo>', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def editar_articulo(id_articulo):
    db = get_db()
    articulo = db.execute('SELECT * FROM articulos WHERE id_articulo=?', (id_articulo,)).fetchone()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()
    if request.method == 'POST':
        cache.clear()
        imagen_filename = articulo['imagen']
        if 'imagen' in request.files:
            file = request.files['imagen']
            if file and allowed_file(file.filename):
                imagen_filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], imagen_filename))
        db.execute('''
            UPDATE articulos SET num_parte=?, nombre=?, descripcion=?, unidad_medida=?,
                stock_minimo=?, link_compra=?, imagen=?, id_proveedor=?, ubicacion=?, stock_actual=?
            WHERE id_articulo=?
        ''', (
            request.form['num_parte'],
            request.form['nombre'],
            request.form['descripcion'],
            request.form['unidad_medida'],
            request.form['stock_minimo'],
            request.form['link_compra'],
            imagen_filename,
            request.form['id_proveedor'],
            request.form.get('ubicacion') or None,
            request.form.get('stock_actual', 0),
            id_articulo
        ))
        db.commit()
        db.close()
        flash('Artículo actualizado correctamente.', 'success')
        return redirect(url_for('articulos'))
    db.close()
    return render_template('editar_articulo.html', articulo=articulo, proveedores=proveedores)

@app.route('/articulos/desactivar/<int:id_articulo>', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def desactivar_articulo(id_articulo):
    if request.method == 'POST':
        cache.clear()
        password = request.form.get('password')
        db = get_db()
        usuario = db.execute(
            'SELECT * FROM usuarios WHERE id_usuario=%s',
            (session['usuario_id'],)
        ).fetchone()
        if not check_password_hash(usuario['password_hash'], password):
            flash('Contraseña incorrecta.', 'error')
            db.close()
            return redirect(url_for('articulos'))
        db.execute('UPDATE articulos SET activo=0 WHERE id_articulo=%s', (id_articulo,))
        db.commit()
        db.close()
        flash('Artículo desactivado correctamente.', 'success')
        return redirect(url_for('articulos'))
    db = get_db()
    articulo = db.execute(
        'SELECT * FROM articulos WHERE id_articulo=%s', (id_articulo,)
    ).fetchone()
    db.close()
    return render_template('confirmar_desactivar.html', articulo=articulo)

# ─── ALERTAS ──────────────────────────────────────────────────────────────────

@app.route('/alertas')
@login_requerido
def alertas():
    db = get_db()
    rol = session.get('rol')
    id_proveedor = session.get('id_proveedor')
    if rol == 'admin':
        lista = db.execute('''
            SELECT a.*, art.nombre as nombre_articulo, art.stock_actual,
                   art.stock_minimo, art.link_compra, p.nombre as nombre_proveedor
            FROM alertas a
            JOIN articulos art ON a.id_articulo = art.id_articulo
            JOIN proveedores p ON art.id_proveedor = p.id_proveedor
            WHERE a.resuelta = 0 ORDER BY a.fecha_hora DESC
        ''').fetchall()
    else:
        lista = db.execute('''
            SELECT a.*, art.nombre as nombre_articulo, art.stock_actual,
                   art.stock_minimo, art.link_compra, p.nombre as nombre_proveedor
            FROM alertas a
            JOIN articulos art ON a.id_articulo = art.id_articulo
            JOIN proveedores p ON art.id_proveedor = p.id_proveedor
            WHERE a.resuelta = 0 AND art.id_proveedor = ?
            ORDER BY a.fecha_hora DESC
        ''', (id_proveedor,)).fetchall()
    db.close()
    return render_template('alertas.html', alertas=lista)

@app.route('/alertas/resolver/<int:id_alerta>')
@login_requerido
def resolver_alerta(id_alerta):
    db = get_db()
    db.execute('UPDATE alertas SET resuelta=1 WHERE id_alerta=?', (id_alerta,))
    db.commit()
    db.close()
    flash('Alerta marcada como resuelta.', 'success')
    return redirect(url_for('alertas'))

# ─── USUARIOS ─────────────────────────────────────────────────────────────────

@app.route('/usuarios')
@login_requerido
@solo_admin
def usuarios():
    db = get_db()
    lista = db.execute('''
        SELECT u.*, p.nombre as nombre_proveedor
        FROM usuarios u
        LEFT JOIN proveedores p ON u.id_proveedor = p.id_proveedor
        WHERE u.activo = 1 ORDER BY u.nombre
    ''').fetchall()
    db.close()
    return render_template('usuarios.html', usuarios=lista)

@app.route('/usuarios/nuevo', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def nuevo_usuario():
    db = get_db()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()
    if request.method == 'POST':
        username = request.form['username']
        existente = db.execute('SELECT * FROM usuarios WHERE username=?', (username,)).fetchone()
        if existente:
            flash('Ese nombre de usuario ya existe.', 'error')
            return render_template('nuevo_usuario.html', proveedores=proveedores)
        id_proveedor = request.form.get('id_proveedor') or None
        db.execute('''
            INSERT INTO usuarios (nombre, username, password_hash, rol, id_proveedor, email)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            request.form['nombre'],
            username,
            generate_password_hash(request.form['password']),
            request.form['rol'],
            id_proveedor,
            request.form.get('email') or None
        ))
        db.commit()
        db.close()
        flash('Usuario creado correctamente.', 'success')
        return redirect(url_for('usuarios'))
    db.close()
    return render_template('nuevo_usuario.html', proveedores=proveedores)

@app.route('/usuarios/editar/<int:id_usuario>', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def editar_usuario(id_usuario):
    db = get_db()
    usuario = db.execute('SELECT * FROM usuarios WHERE id_usuario=?', (id_usuario,)).fetchone()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()
    if request.method == 'POST':
        id_proveedor = request.form.get('id_proveedor') or None
        if request.form['password']:
            password_hash = generate_password_hash(request.form['password'])
        else:
            password_hash = usuario['password_hash']
        db.execute('''
            UPDATE usuarios SET nombre=?, username=?, password_hash=?, rol=?, id_proveedor=?, email=?
            WHERE id_usuario=?
        ''', (
            request.form['nombre'],
            request.form['username'],
            password_hash,
            request.form['rol'],
            id_proveedor,
            request.form.get('email') or None,
            id_usuario
        ))
        db.commit()
        db.close()
        flash('Usuario actualizado correctamente.', 'success')
        return redirect(url_for('usuarios'))
    db.close()
    return render_template('editar_usuario.html', usuario=usuario, proveedores=proveedores)

@app.route('/usuarios/desactivar/<int:id_usuario>')
@login_requerido
@solo_admin
def desactivar_usuario(id_usuario):
    if id_usuario == session.get('usuario_id'):
        flash('No puedes desactivar tu propia cuenta.', 'error')
        return redirect(url_for('usuarios'))
    db = get_db()
    db.execute('UPDATE usuarios SET activo=0 WHERE id_usuario=?', (id_usuario,))
    db.commit()
    db.close()
    flash('Usuario desactivado.', 'success')
    return redirect(url_for('usuarios'))

# ─── ESTADÍSTICAS ─────────────────────────────────────────────────────────────

@app.route('/estadisticas')
@login_requerido
def estadisticas():
    db = get_db()
    rol = session.get('rol')
    id_proveedor = session.get('id_proveedor')

    if rol == 'admin':
        top_articulos = db.execute('''
            SELECT art.nombre, art.unidad_medida,
                   SUM(CASE WHEN m.tipo='SALIDA' THEN m.cantidad ELSE 0 END) as total_salidas,
                   SUM(CASE WHEN m.tipo='ENTRADA' THEN m.cantidad ELSE 0 END) as total_entradas
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            GROUP BY art.id_articulo
            ORDER BY total_salidas DESC LIMIT 10
        ''').fetchall()

        consumo_mensual = db.execute('''
            SELECT to_char(fecha_hora, 'MM/YYYY') as mes,
                   SUM(CASE WHEN tipo='SALIDA' THEN cantidad ELSE 0 END) as salidas,
                   SUM(CASE WHEN tipo='ENTRADA' THEN cantidad ELSE 0 END) as entradas
            FROM movimientos
            GROUP BY to_char(fecha_hora, 'MM/YYYY') ORDER BY MIN(fecha_hora) ASC LIMIT 12
        ''').fetchall()

        por_proveedor = db.execute('''
            SELECT p.nombre as proveedor,
                   SUM(CASE WHEN m.tipo='SALIDA' THEN m.cantidad ELSE 0 END) as total_salidas
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            JOIN proveedores p ON art.id_proveedor = p.id_proveedor
            GROUP BY p.id_proveedor ORDER BY total_salidas DESC
        ''').fetchall()
    else:
        top_articulos = db.execute('''
            SELECT art.nombre, art.unidad_medida,
                   SUM(CASE WHEN m.tipo='SALIDA' THEN m.cantidad ELSE 0 END) as total_salidas,
                   SUM(CASE WHEN m.tipo='ENTRADA' THEN m.cantidad ELSE 0 END) as total_entradas
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            WHERE art.id_proveedor = ?
            GROUP BY art.id_articulo
            ORDER BY total_salidas DESC LIMIT 10
        ''', (id_proveedor,)).fetchall()

        consumo_mensual = db.execute('''
            SELECT to_char(fecha_hora, 'MM/YYYY') as mes,
                   SUM(CASE WHEN tipo='SALIDA' THEN cantidad ELSE 0 END) as salidas,
                   SUM(CASE WHEN tipo='ENTRADA' THEN cantidad ELSE 0 END) as entradas
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            WHERE art.id_proveedor = ?
            GROUP BY to_char(fecha_hora, 'MM/YYYY') ORDER BY MIN(fecha_hora) ASC LIMIT 12
        ''', (id_proveedor,)).fetchall()

        por_proveedor = []

    # Actualizar materiales críticos automáticamente
    if rol == 'admin':
        db.execute('UPDATE articulos SET es_critico = 0')
        criticos = db.execute('''
            SELECT art.id_articulo
            FROM movimientos m
            JOIN articulos art ON m.id_articulo = art.id_articulo
            WHERE m.tipo = 'SALIDA'
            GROUP BY art.id_articulo
            ORDER BY SUM(m.cantidad) DESC LIMIT 5
        ''').fetchall()
        for c in criticos:
            db.execute('UPDATE articulos SET es_critico = 1 WHERE id_articulo = ?', (c['id_articulo'],))
        db.commit()

    db.close()
    return render_template('estadisticas.html',
        top_articulos=top_articulos,
        consumo_mensual=consumo_mensual,
        por_proveedor=por_proveedor
    )

# ─── EXPORTACIÓN ──────────────────────────────────────────────────────────────

@app.route('/exportar')
@login_requerido
def exportar():
    db = get_db()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()
    db.close()
    return render_template('exportar.html', proveedores=proveedores)

@app.route('/exportar/inventario')
@login_requerido
def exportar_inventario():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    db = get_db()
    rol = session.get('rol')
    id_proveedor_filtro = request.args.get('id_proveedor')
    id_proveedor_session = session.get('id_proveedor')

    if rol == 'admin' and id_proveedor_filtro:
        articulos = db.execute('''
            SELECT a.*, p.nombre as nombre_proveedor
            FROM articulos a JOIN proveedores p ON a.id_proveedor = p.id_proveedor
            WHERE a.activo=1 AND a.id_proveedor=? ORDER BY p.nombre, a.nombre
        ''', (id_proveedor_filtro,)).fetchall()
    elif rol == 'admin':
        articulos = db.execute('''
            SELECT a.*, p.nombre as nombre_proveedor
            FROM articulos a JOIN proveedores p ON a.id_proveedor = p.id_proveedor
            WHERE a.activo=1 ORDER BY p.nombre, a.nombre
        ''').fetchall()
    else:
        articulos = db.execute('''
            SELECT a.*, p.nombre as nombre_proveedor
            FROM articulos a JOIN proveedores p ON a.id_proveedor = p.id_proveedor
            WHERE a.activo=1 AND a.id_proveedor=? ORDER BY a.nombre
        ''', (id_proveedor_session,)).fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos
    header_fill = PatternFill("solid", fgColor="0B4F6C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EBF5FB")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center = Alignment(horizontal='center', vertical='center')

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = 'INVENTARIO GENERAL - ALMACÉN MEDLINE'
    ws['A1'].font = Font(bold=True, size=14, color="0B4F6C")
    ws['A1'].alignment = center

    ws.merge_cells('A2:I2')
    ws['A2'] = f'Generado el: {datetime.now().strftime("%m/%d/%Y %I:%M %p")}'
    ws['A2'].alignment = center
    ws['A2'].font = Font(size=10, color="888888")

    # Encabezados
    headers = ['#', 'Num. Parte', 'Nombre', 'Proveedor', 'Descripción',
               'Unidad', 'Stock Actual', 'Stock Mínimo', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Datos
    for row_idx, art in enumerate(articulos, 1):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        if art['stock_actual'] <= art['stock_minimo']:
            estado = 'STOCK BAJO'
            estado_fill = PatternFill("solid", fgColor="FADBD8")
        elif art['stock_actual'] <= art['stock_minimo'] * 2:
            estado = 'PRECAUCIÓN'
            estado_fill = PatternFill("solid", fgColor="FDEBD0")
        else:
            estado = 'OK'
            estado_fill = PatternFill("solid", fgColor="D5F5E3")

        row_data = [
            row_idx,
            art['num_parte'],
            art['nombre'],
            art['nombre_proveedor'],
            art['descripcion'] or '',
            art['unidad_medida'],
            art['stock_actual'],
            art['stock_minimo'],
            estado
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx + 4, column=col, value=value)
            cell.border = border
            cell.alignment = center
            if col == 9:
                cell.fill = estado_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = fill

    # Ancho de columnas
    anchos = [5, 14, 28, 20, 30, 12, 14, 14, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventario_{datetime.now().strftime("%m%d%Y")}.xlsx'
    )

@app.route('/exportar/movimientos')
@login_requerido
def exportar_movimientos():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    db = get_db()
    rol = session.get('rol')
    id_proveedor = session.get('id_proveedor')
    fecha_ini = request.args.get('fecha_ini', '')
    fecha_fin = request.args.get('fecha_fin', '')

    query = '''
        SELECT m.*, art.nombre as nombre_articulo, art.num_parte,
               p.nombre as nombre_proveedor, u.nombre as nombre_usuario
        FROM movimientos m
        JOIN articulos art ON m.id_articulo = art.id_articulo
        JOIN proveedores p ON art.id_proveedor = p.id_proveedor
        LEFT JOIN usuarios u ON m.id_usuario = u.id_usuario
    '''
    params = []
    conditions = []

    if rol != 'admin':
        conditions.append('art.id_proveedor = ?')
        params.append(id_proveedor)
    if fecha_ini:
        conditions.append('DATE(m.fecha_hora) >= ?')
        params.append(fecha_ini)
    if fecha_fin:
        conditions.append('DATE(m.fecha_hora) <= ?')
        params.append(fecha_fin)

    if conditions:
        query += ' WHERE ' + ' AND '.join(conditions)
    query += ' ORDER BY m.fecha_hora DESC'

    movimientos = db.execute(query, params).fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    header_fill = PatternFill("solid", fgColor="0B4F6C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:H1')
    ws['A1'] = 'HISTORIAL DE MOVIMIENTOS - ALMACÉN MEDLINE'
    ws['A1'].font = Font(bold=True, size=14, color="0B4F6C")
    ws['A1'].alignment = center

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generado el: {datetime.now().strftime("%m/%d/%Y %I:%M %p")}'
    ws['A2'].alignment = center
    ws['A2'].font = Font(size=10, color="888888")

    headers = ['#', 'Fecha', 'Artículo', 'Num. Parte', 'Proveedor', 'Tipo', 'Cantidad', 'Usuario']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_idx, mov in enumerate(movimientos, 1):
        tipo_fill = PatternFill("solid", fgColor="D5F5E3") if mov['tipo'] == 'ENTRADA' else PatternFill("solid", fgColor="FADBD8")
        row_data = [
            row_idx,
            mov['fecha_hora'],
            mov['nombre_articulo'],
            mov['num_parte'],
            mov['nombre_proveedor'],
            mov['tipo'],
            mov['cantidad'],
            mov['nombre_usuario'] or 'Kiosco'
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx + 4, column=col, value=value)
            cell.border = border
            cell.alignment = center
            if col == 6:
                cell.fill = tipo_fill
                cell.font = Font(bold=True)

    anchos = [5, 20, 28, 14, 20, 12, 12, 20]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'movimientos_{datetime.now().strftime("%m%d%Y")}.xlsx'
    )

@app.route('/exportar/por_proveedor')
@login_requerido
@solo_admin
def exportar_por_proveedor():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    db = get_db()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="0B4F6C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center = Alignment(horizontal='center', vertical='center')

    for prov in proveedores:
        ws = wb.create_sheet(title=prov['nombre'][:31])
        articulos = db.execute('''
            SELECT * FROM articulos WHERE id_proveedor=? AND activo=1 ORDER BY nombre
        ''', (prov['id_proveedor'],)).fetchall()

        ws.merge_cells('A1:H1')
        ws['A1'] = f'INVENTARIO - {prov["nombre"].upper()}'
        ws['A1'].font = Font(bold=True, size=13, color="0B4F6C")
        ws['A1'].alignment = center

        ws.merge_cells('A2:H2')
        ws['A2'] = f'Generado el: {datetime.now().strftime("%m/%d/%Y %I:%M %p")}'
        ws['A2'].alignment = center
        ws['A2'].font = Font(size=10, color="888888")

        headers = ['#', 'Num. Parte', 'Nombre', 'Descripción', 'Unidad', 'Stock Actual', 'Stock Mínimo', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row_idx, art in enumerate(articulos, 1):
            if art['stock_actual'] <= art['stock_minimo']:
                estado = 'STOCK BAJO'
                estado_fill = PatternFill("solid", fgColor="FADBD8")
            elif art['stock_actual'] <= art['stock_minimo'] * 2:
                estado = 'PRECAUCIÓN'
                estado_fill = PatternFill("solid", fgColor="FDEBD0")
            else:
                estado = 'OK'
                estado_fill = PatternFill("solid", fgColor="D5F5E3")

            row_data = [row_idx, art['num_parte'], art['nombre'],
                       art['descripcion'] or '', art['unidad_medida'],
                       art['stock_actual'], art['stock_minimo'], estado]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 4, column=col, value=value)
                cell.border = border
                cell.alignment = center
                if col == 8:
                    cell.fill = estado_fill
                    cell.font = Font(bold=True)

        anchos = [5, 14, 28, 30, 12, 14, 14, 14]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

    db.close()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventario_proveedores_{datetime.now().strftime("%m%d%Y")}.xlsx'
    )

@app.route('/exportar/sql')
@login_requerido
@solo_admin
def exportar_sql():
    import io
    db = get_db()
    output = io.StringIO()
    
    output.write('-- ================================================\n')
    output.write('-- SISTEMA DE ALMACÉN MEDLINE\n')
    output.write(f'-- Generado el: {datetime.now().strftime("%m/%d/%Y %I:%M %p")}\n')
    output.write('-- ================================================\n\n')

    output.write('-- Puedes correr este archivo en MySQL, MariaDB o SQL Server\n\n')

    # Estructura y datos de cada tabla
    tablas = ['proveedores', 'usuarios', 'articulos', 'movimientos', 'alertas', 'ordenes_compra']

    create_statements = {
        'proveedores': '''CREATE TABLE IF NOT EXISTS proveedores (
    id_proveedor INT AUTO_INCREMENT PRIMARY KEY,
    nombre VARCHAR(255) NOT NULL,
    contacto VARCHAR(255),
    area VARCHAR(255),
    logo VARCHAR(255),
    activo TINYINT DEFAULT 1
);\n''',
        'usuarios': '''CREATE TABLE IF NOT EXISTS usuarios (
    id_usuario INT AUTO_INCREMENT PRIMARY KEY,
    nombre VARCHAR(255) NOT NULL,
    username VARCHAR(255) UNIQUE NOT NULL,
    password_hash VARCHAR(512) NOT NULL,
    rol VARCHAR(50) NOT NULL,
    id_proveedor INT,
    email VARCHAR(255),
    activo TINYINT DEFAULT 1,
    FOREIGN KEY (id_proveedor) REFERENCES proveedores(id_proveedor)
);\n''',
        'articulos': '''CREATE TABLE IF NOT EXISTS articulos (
    id_articulo INT AUTO_INCREMENT PRIMARY KEY,
    num_parte VARCHAR(255) NOT NULL,
    nombre VARCHAR(255) NOT NULL,
    descripcion TEXT,
    unidad_medida VARCHAR(50) DEFAULT 'piezas',
    stock_actual INT DEFAULT 0,
    stock_minimo INT DEFAULT 5,
    es_critico TINYINT DEFAULT 0,
    codigo_qr VARCHAR(255) UNIQUE,
    link_compra TEXT,
    imagen VARCHAR(255),
    id_proveedor INT NOT NULL,
    activo TINYINT DEFAULT 1,
    FOREIGN KEY (id_proveedor) REFERENCES proveedores(id_proveedor)
);\n''',
        'movimientos': '''CREATE TABLE IF NOT EXISTS movimientos (
    id_movimiento INT AUTO_INCREMENT PRIMARY KEY,
    id_articulo INT NOT NULL,
    id_usuario INT,
    tipo VARCHAR(10) NOT NULL,
    cantidad INT NOT NULL,
    fecha_hora DATETIME DEFAULT CURRENT_TIMESTAMP,
    observaciones TEXT,
    modo_kiosco TINYINT DEFAULT 0,
    FOREIGN KEY (id_articulo) REFERENCES articulos(id_articulo),
    FOREIGN KEY (id_usuario) REFERENCES usuarios(id_usuario)
);\n''',
        'alertas': '''CREATE TABLE IF NOT EXISTS alertas (
    id_alerta INT AUTO_INCREMENT PRIMARY KEY,
    id_articulo INT NOT NULL,
    tipo_alerta VARCHAR(50) NOT NULL,
    fecha_hora DATETIME DEFAULT CURRENT_TIMESTAMP,
    vista TINYINT DEFAULT 0,
    resuelta TINYINT DEFAULT 0,
    id_usuario_dest INT,
    FOREIGN KEY (id_articulo) REFERENCES articulos(id_articulo)
);\n''',
        'ordenes_compra': '''CREATE TABLE IF NOT EXISTS ordenes_compra (
    id_orden INT AUTO_INCREMENT PRIMARY KEY,
    id_articulo INT NOT NULL,
    cantidad_sugerida INT NOT NULL,
    fecha_hora DATETIME DEFAULT CURRENT_TIMESTAMP,
    estado VARCHAR(50) DEFAULT 'pendiente',
    notas TEXT,
    id_usuario INT,
    FOREIGN KEY (id_articulo) REFERENCES articulos(id_articulo)
);\n'''
    }

    for tabla in tablas:
        output.write(f'-- ── Tabla: {tabla} ──\n')
        output.write(create_statements[tabla])
        output.write('\n')

        filas = db.execute(f'SELECT * FROM {tabla}').fetchall()
        if filas:
            cols = filas[0].keys()
            for fila in filas:
                valores = []
                for v in fila:
                    if v is None:
                        valores.append('NULL')
                    elif isinstance(v, str):
                        v_escaped = v.replace("'", "''")
                        valores.append(f"'{v_escaped}'")
                    else:
                        valores.append(str(v))
                cols_str = ', '.join(cols)
                vals_str = ', '.join(valores)
                output.write(f'INSERT INTO {tabla} ({cols_str}) VALUES ({vals_str});\n')
        output.write('\n')

    db.close()

    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='text/plain',
        headers={
            'Content-Disposition': f'attachment; filename=almacen_medline_{datetime.now().strftime("%m%d%Y")}.sql'
        }
    )

# ─── ÓRDENES DE COMPRA ────────────────────────────────────────────────────────

@app.route('/orden_compra/<int:id_articulo>')
@login_requerido
def orden_compra(id_articulo):
    db = get_db()
    articulo = db.execute('''
        SELECT a.*, p.nombre as nombre_proveedor, p.contacto as contacto_proveedor
        FROM articulos a JOIN proveedores p ON a.id_proveedor = p.id_proveedor
        WHERE a.id_articulo = ?
    ''', (id_articulo,)).fetchone()
    db.close()
    return render_template('orden_compra.html', articulo=articulo,
                           fecha=datetime.now().strftime('%m/%d/%Y'),
                           folio=f'OC-{id_articulo}-{datetime.now().strftime("%m%d%Y")}')

@app.route('/orden_compra/pdf/<int:id_articulo>')
@login_requerido
def orden_compra_pdf(id_articulo):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    import io

    db = get_db()
    articulo = db.execute('''
        SELECT a.*, p.nombre as nombre_proveedor, p.contacto as contacto_proveedor
        FROM articulos a JOIN proveedores p ON a.id_proveedor = p.id_proveedor
        WHERE a.id_articulo = ?
    ''', (id_articulo,)).fetchone()

    cantidad_sugerida = (articulo['stock_minimo'] * 2) - articulo['stock_actual']
    if cantidad_sugerida < 1:
        cantidad_sugerida = articulo['stock_minimo']

    db.execute('''
        INSERT INTO ordenes_compra (id_articulo, cantidad_sugerida, id_usuario)
        VALUES (?, ?, ?)
    ''', (id_articulo, cantidad_sugerida, session.get('usuario_id')))
    db.commit()
    db.close()

    folio = f'OC-{id_articulo}-{datetime.now().strftime("%m%d%Y")}'
    fecha = datetime.now().strftime('%m/%d/%Y')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)

    styles = getSampleStyleSheet()
    elementos = []

    # Header
    header_style = ParagraphStyle('header', fontSize=18, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#0B4F6C'), spaceAfter=4)
    sub_style = ParagraphStyle('sub', fontSize=10, textColor=colors.HexColor('#888888'), spaceAfter=20)
    label_style = ParagraphStyle('label', fontSize=10, fontName='Helvetica-Bold',
                                 textColor=colors.HexColor('#0B4F6C'))
    normal_style = ParagraphStyle('normal', fontSize=10, spaceAfter=6)

    elementos.append(Paragraph('ORDEN DE COMPRA', header_style))
    elementos.append(Paragraph('Sistema de Almacén - Medline', sub_style))

    # Info folio y fecha
    info_data = [
        ['Folio:', folio, 'Fecha:', fecha],
        ['Solicitante:', session.get('nombre'), 'Proveedor:', articulo['nombre_proveedor']],
    ]
    info_table = Table(info_data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 2.5*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#0B4F6C')),
        ('TEXTCOLOR', (2,0), (2,-1), colors.HexColor('#0B4F6C')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,-1), (-1,-1), 1, colors.HexColor('#CCCCCC')),
    ]))
    elementos.append(info_table)
    elementos.append(Spacer(1, 0.3*inch))

    # Tabla del artículo
    elementos.append(Paragraph('DETALLE DEL ARTÍCULO', label_style))
    elementos.append(Spacer(1, 0.1*inch))

    art_data = [
        ['Campo', 'Detalle'],
        ['Nombre del artículo', articulo['nombre']],
        ['Número de parte', articulo['num_parte']],
        ['Descripción', articulo['descripcion'] or 'N/A'],
        ['Unidad de medida', articulo['unidad_medida']],
        ['Stock actual', str(articulo['stock_actual'])],
        ['Stock mínimo', str(articulo['stock_minimo'])],
        ['Cantidad solicitada', str(cantidad_sugerida)],
        ['Link de compra', articulo['link_compra'] or 'N/A'],
    ]

    art_table = Table(art_data, colWidths=[2.5*inch, 5*inch])
    art_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0B4F6C')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0,1), (0,-1), colors.HexColor('#0B4F6C')),
        ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#EBF5FB')),
        ('BACKGROUND', (0,4), (-1,4), colors.HexColor('#EBF5FB')),
        ('BACKGROUND', (0,6), (-1,6), colors.HexColor('#EBF5FB')),
        ('BACKGROUND', (0,8), (-1,8), colors.HexColor('#EBF5FB')),
        ('BACKGROUND', (0,7), (-1,7), colors.HexColor('#FDEBD0')),
        ('FONTNAME', (0,7), (-1,7), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUND', (0,0), (-1,0), colors.HexColor('#0B4F6C')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elementos.append(art_table)
    elementos.append(Spacer(1, 0.4*inch))

    # Firmas
    elementos.append(Paragraph('AUTORIZACIONES', label_style))
    elementos.append(Spacer(1, 0.2*inch))

    firmas_data = [
        ['_________________________', '_________________________', '_________________________'],
        ['Solicitante', 'Jefe de Almacén', 'Finanzas'],
        [session.get('nombre'), 'Nombre y firma', 'Nombre y firma'],
    ]
    firmas_table = Table(firmas_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
    firmas_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0,1), (-1,1), colors.HexColor('#0B4F6C')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
    ]))
    elementos.append(firmas_table)

    # Footer
    elementos.append(Spacer(1, 0.3*inch))
    footer_style = ParagraphStyle('footer', fontSize=8, textColor=colors.HexColor('#AAAAAA'),
                                  alignment=1)
    elementos.append(Paragraph(
        f'Documento generado automáticamente por el Sistema de Almacén Medline · {fecha}',
        footer_style
    ))

    doc.build(elementos)
    buffer.seek(0)

    from flask import send_file
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'orden_compra_{folio}.pdf'
    )

# ─── IMPORTACIÓN ──────────────────────────────────────────────────────────────

@app.route('/importar', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def importar():
    db = get_db()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall()
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('importar'))
        archivo = request.files['archivo']
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            flash('Solo se permiten archivos Excel (.xlsx o .xls).', 'error')
            return redirect(url_for('importar'))

        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(archivo.read()))
        ws = wb.active

        modo = request.form.get('modo')
        id_proveedor_fijo = request.form.get('id_proveedor') or None
        exitosos = 0
        errores = 0
        errores_detalle = []

        # Construir mapa de nombre proveedor -> id
        mapa_proveedores = {}
        for p in proveedores:
            mapa_proveedores[p['nombre'].strip().lower()] = p['id_proveedor']

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0] and not row[1]:
                    continue

                if modo == 'general':
                    # Columna H tiene el nombre del proveedor
                    nombre_prov = str(row[7]).strip().lower() if len(row) > 7 and row[7] else None
                    if not nombre_prov or nombre_prov not in mapa_proveedores:
                        errores += 1
                        errores_detalle.append(f'Fila {row_idx}: Proveedor "{row[7]}" no encontrado.')
                        continue
                    id_proveedor = mapa_proveedores[nombre_prov]
                else:
                    id_proveedor = id_proveedor_fijo

                num_parte = str(row[0]).strip() if row[0] else None
                nombre = str(row[1]).strip() if row[1] else None
                descripcion = str(row[2]).strip() if row[2] else ''
                unidad_medida = str(row[3]).strip() if row[3] else 'piezas'
                stock_inicial = int(row[4]) if row[4] else 0
                stock_minimo = int(row[5]) if row[5] else 5
                link_compra = str(row[6]).strip() if len(row) > 6 and row[6] else None

                if not num_parte or not nombre:
                    errores += 1
                    errores_detalle.append(f'Fila {row_idx}: Num. parte y nombre son obligatorios.')
                    continue

                existente = db.execute(
                    'SELECT * FROM articulos WHERE num_parte=? AND id_proveedor=?',
                    (num_parte, id_proveedor)
                ).fetchone()

                if existente:
                    db.execute('''
                        UPDATE articulos SET nombre=?, descripcion=?, unidad_medida=?,
                            stock_minimo=?, link_compra=?
                        WHERE num_parte=? AND id_proveedor=?
                    ''', (nombre, descripcion, unidad_medida, stock_minimo,
                          link_compra, num_parte, id_proveedor))
                else:
                    db.execute('''
                        INSERT INTO articulos (num_parte, nombre, descripcion, unidad_medida,
                            stock_actual, stock_minimo, codigo_qr, link_compra, id_proveedor)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (num_parte, nombre, descripcion, unidad_medida,
                          stock_inicial, stock_minimo, num_parte, link_compra, id_proveedor))
                exitosos += 1
            except Exception as e:
                errores += 1
                errores_detalle.append(f'Fila {row_idx}: {str(e)}')

        db.commit()
        db.close()

        if exitosos > 0:
            flash(f'✅ {exitosos} artículos importados correctamente.', 'success')
        if errores > 0:
            flash(f'⚠️ {errores} filas con error: {" | ".join(errores_detalle[:3])}', 'error')

        return redirect(url_for('articulos'))

    db.close()
    return render_template('importar.html', proveedores=proveedores)

@app.route('/importar/plantilla')
@login_requerido
@solo_admin
def descargar_plantilla():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    header_fill = PatternFill("solid", fgColor="0B4F6C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center')

    headers = ['Num. Parte', 'Nombre', 'Descripcion', 'Unidad Medida',
               'Stock Inicial', 'Stock Minimo', 'Link Compra', 'Proveedor (solo modo general)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Filas de ejemplo
    ejemplos = [
        ['MP-001', 'Guantes de nitrilo', 'Guantes desechables talla M', 'cajas', 10, 5, 'https://ejemplo.com'],
        ['MP-002', 'Cubrebocas N95', 'Mascarilla de protección', 'piezas', 50, 20, ''],
        ['MP-003', 'Gel antibacterial', 'Botella 1 litro', 'litros', 8, 3, ''],
    ]
    for row_idx, ejemplo in enumerate(ejemplos, 2):
        for col, value in enumerate(ejemplo, 1):
            ws.cell(row=row_idx, column=col, value=value)

    anchos = [14, 28, 30, 16, 14, 14, 35, 25]
    from openpyxl.utils import get_column_letter
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_inventario.xlsx'
    )

@app.route('/articulos/reasignar', methods=['POST'])
@login_requerido
@solo_admin
def reasignar_articulos():
    ids = request.form.getlist('articulos_sel')
    id_proveedor_nuevo = request.form.get('id_proveedor_nuevo')
    if not ids or not id_proveedor_nuevo:
        flash('Selecciona al menos un artículo y un proveedor.', 'error')
        return redirect(url_for('articulos'))
    db = get_db()
    for id_art in ids:
        db.execute('UPDATE articulos SET id_proveedor=? WHERE id_articulo=?',
                   (id_proveedor_nuevo, id_art))
    db.commit()
    db.close()
    flash(f'✅ {len(ids)} artículo(s) reasignado(s) correctamente.', 'success')
    return redirect(url_for('articulos'))

# ─── DETALLE ARTÍCULO ─────────────────────────────────────────────────────────

@app.route('/articulos/<int:id_articulo>')
@login_requerido
def detalle_articulo(id_articulo):
    db = get_db()
    articulo = db.execute('''
        SELECT a.*, p.nombre as nombre_proveedor
        FROM articulos a JOIN proveedores p ON a.id_proveedor = p.id_proveedor
        WHERE a.id_articulo = ?
    ''', (id_articulo,)).fetchone()

    movimientos = db.execute('''
        SELECT m.*, u.nombre as nombre_usuario
        FROM movimientos m
        LEFT JOIN usuarios u ON m.id_usuario = u.id_usuario
        WHERE m.id_articulo = ?
        ORDER BY m.fecha_hora DESC LIMIT 20
    ''', (id_articulo,)).fetchall()

    total_entradas = db.execute(
        "SELECT SUM(cantidad) FROM movimientos WHERE id_articulo=? AND tipo='ENTRADA'",
        (id_articulo,)
    ).fetchone()[0] or 0

    total_salidas = db.execute(
        "SELECT SUM(cantidad) FROM movimientos WHERE id_articulo=? AND tipo='SALIDA'",
        (id_articulo,)
    ).fetchone()[0] or 0

    # Generar código de barras
    import barcode
    from barcode.writer import ImageWriter
    import io, base64
    try:
        code128 = barcode.get('code128', str(articulo['num_parte']), writer=ImageWriter())
        buffer = io.BytesIO()
        code128.write(buffer, options={
            'module_height': 10,
            'module_width': 0.3,
            'quiet_zone': 2,
            'font_size': 8,
            'text_distance': 2,
            'background': 'white',
            'foreground': 'black',
            'write_text': False
        })
        barcode_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    except Exception:
        barcode_b64 = None

    db.close()
    return render_template('detalle_articulo.html',
        articulo=articulo,
        movimientos=movimientos,
        total_entradas=total_entradas,
        total_salidas=total_salidas,
        barcode_b64=barcode_b64
    )

# ─── PERFIL ───────────────────────────────────────────────────────────────────

@app.route('/perfil', methods=['GET', 'POST'])
@login_requerido
def perfil():
    db = get_db()
    usuario = db.execute(
        'SELECT * FROM usuarios WHERE id_usuario=?', (session['usuario_id'],)
    ).fetchone()
    if request.method == 'POST':
        nombre = request.form['nombre']
        email = request.form.get('email') or None
        password_nueva = request.form.get('password_nueva')
        password_actual = request.form.get('password_actual')
        if password_nueva:
            if not check_password_hash(usuario['password_hash'], password_actual):
                flash('La contraseña actual es incorrecta.', 'error')
                db.close()
                return render_template('perfil.html', usuario=usuario)
            db.execute('''
                UPDATE usuarios SET nombre=%s, email=%s, password_hash=%s, area=%s, puesto=%s WHERE id_usuario=%s
            ''', (nombre, email, generate_password_hash(password_nueva), area, puesto, session['usuario_id']))
        else:
         area = request.form.get('area') or None
        puesto = request.form.get('puesto') or None
        db.execute('''
            UPDATE usuarios SET nombre=%s, email=%s, area=%s, puesto=%s WHERE id_usuario=%s
        ''', (nombre, email, area, puesto, session['usuario_id']))
        db.commit()
        session['nombre'] = nombre
        flash('Perfil actualizado correctamente.', 'success')
        db.close()
        return redirect(url_for('perfil'))
    proveedor_logo = None
    if usuario['rol'] == 'proveedor' and usuario['id_proveedor']:
        prov = db.execute('SELECT logo FROM proveedores WHERE id_proveedor=?',
                         (usuario['id_proveedor'],)).fetchone()
        if prov and prov['logo']:
            proveedor_logo = prov['logo']
    db.close()
    return render_template('perfil.html', usuario=usuario, proveedor_logo=proveedor_logo)

@app.route('/api/mi_perfil')
@login_requerido
def api_mi_perfil():
    db = get_db()
    usuario = db.execute(
        'SELECT nombre, area, puesto FROM usuarios WHERE id_usuario=%s',
        (session['usuario_id'],)
    ).fetchone()
    db.close()
    return jsonify({
        'nombre': usuario['nombre'] if usuario else '',
        'area': usuario['area'] if usuario else '',
        'puesto': usuario['puesto'] if usuario else ''
    })

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

@app.route('/configuracion', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def configuracion():
    config_path = 'config.json'
    import json

    # Cargar config actual
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            config = json.load(f)
    else:
        config = {
            'nombre_almacen': 'Almacén Medline',
            'stock_minimo_global': 5,
            'correo_alertas': '',
            'dias_critico': 30
        }

    if request.method == 'POST':
        config['nombre_almacen'] = request.form.get('nombre_almacen', 'Almacén Medline')
        config['stock_minimo_global'] = int(request.form.get('stock_minimo_global', 5))
        config['correo_alertas'] = request.form.get('correo_alertas', '')
        config['dias_critico'] = int(request.form.get('dias_critico', 30))
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=2)
        flash('Configuración guardada correctamente.', 'success')
        return redirect(url_for('configuracion'))

    return render_template('configuracion.html', config=config)

# ─── HISTORIAL ────────────────────────────────────────────────────────────────

@app.route('/historial')
@login_requerido
def historial():
    db = get_db()
    rol = session.get('rol')
    id_proveedor = session.get('id_proveedor')

    fecha_ini = request.args.get('fecha_ini', '')
    fecha_fin = request.args.get('fecha_fin', '')
    tipo = request.args.get('tipo', '')
    id_prov_filtro = request.args.get('id_proveedor', '')

    query = '''
        SELECT m.*, art.nombre as nombre_articulo, art.num_parte,
               p.nombre as nombre_proveedor, u.nombre as nombre_usuario
        FROM movimientos m
        JOIN articulos art ON m.id_articulo = art.id_articulo
        JOIN proveedores p ON art.id_proveedor = p.id_proveedor
        LEFT JOIN usuarios u ON m.id_usuario = u.id_usuario
    '''
    params = []
    conditions = []

    if rol != 'admin':
        conditions.append('art.id_proveedor = ?')
        params.append(id_proveedor)
    elif id_prov_filtro:
        conditions.append('art.id_proveedor = ?')
        params.append(id_prov_filtro)

    if fecha_ini:
        conditions.append('DATE(m.fecha_hora) >= ?')
        params.append(fecha_ini)
    if fecha_fin:
        conditions.append('DATE(m.fecha_hora) <= ?')
        params.append(fecha_fin)
    if tipo:
        conditions.append('m.tipo = ?')
        params.append(tipo)

    if conditions:
        query += ' WHERE ' + ' AND '.join(conditions)
    query += ' ORDER BY m.fecha_hora DESC'

    movimientos = db.execute(query, params).fetchall()
    proveedores = db.execute('SELECT * FROM proveedores WHERE activo=1').fetchall() if rol == 'admin' else []

    total_entradas = sum(m['cantidad'] for m in movimientos if m['tipo'] == 'ENTRADA')
    total_salidas = sum(m['cantidad'] for m in movimientos if m['tipo'] == 'SALIDA')

    db.close()
    return render_template('historial.html',
        movimientos=movimientos,
        proveedores=proveedores,
        total_entradas=total_entradas,
        total_salidas=total_salidas,
        fecha_ini=fecha_ini,
        fecha_fin=fecha_fin,
        tipo=tipo,
        id_prov_filtro=id_prov_filtro
    )

@app.route('/importar/sql', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def importar_sql():
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('importar_sql'))
        archivo = request.files['archivo']
        if not archivo.filename.endswith('.sql'):
            flash('Solo se permiten archivos .sql', 'error')
            return redirect(url_for('importar_sql'))
        try:
            contenido = archivo.read().decode('utf-8')
            db = get_db()
            sentencias = [s.strip() for s in contenido.split(';') if s.strip()]
            exitosas = 0
            errores = 0
            errores_detalle = []
            for sentencia in sentencias:
                if sentencia.startswith('--') or not sentencia:
                    continue
                try:
                    db.execute(sentencia)
                    exitosas += 1
                except Exception as e:
                    errores += 1
                    errores_detalle.append(str(e)[:100])
            db.commit()
            db.close()
            if exitosas > 0:
                flash(f'✅ {exitosas} sentencias ejecutadas correctamente.', 'success')
            if errores > 0:
                flash(f'⚠️ {errores} sentencias con error: {" | ".join(errores_detalle[:3])}', 'error')
        except Exception as e:
            flash(f'Error al leer el archivo: {str(e)}', 'error')
        return redirect(url_for('importar_sql'))
    return render_template('importar_sql.html')

# ─── BÚSQUEDA GLOBAL ──────────────────────────────────────────────────────────

@app.route('/buscar')
@login_requerido
def buscar():
    q = request.args.get('q', '').strip()
    if not q:
        return redirect(url_for('dashboard'))
    db = get_db()
    rol = session.get('rol')
    id_proveedor = session.get('id_proveedor')

    if rol == 'admin':
        articulos = db.execute('''
            SELECT a.*, p.nombre as nombre_proveedor FROM articulos a
            JOIN proveedores p ON a.id_proveedor = p.id_proveedor
            WHERE a.activo=1 AND (a.nombre LIKE ? OR a.num_parte LIKE ? OR a.descripcion LIKE ?)
        ''', (f'%{q}%', f'%{q}%', f'%{q}%')).fetchall()
        proveedores = db.execute('''
            SELECT * FROM proveedores WHERE activo=1 AND (nombre LIKE ? OR area LIKE ?)
        ''', (f'%{q}%', f'%{q}%')).fetchall()
    else:
        articulos = db.execute('''
            SELECT a.*, p.nombre as nombre_proveedor FROM articulos a
            JOIN proveedores p ON a.id_proveedor = p.id_proveedor
            WHERE a.activo=1 AND a.id_proveedor=?
            AND (a.nombre LIKE ? OR a.num_parte LIKE ? OR a.descripcion LIKE ?)
        ''', (id_proveedor, f'%{q}%', f'%{q}%', f'%{q}%')).fetchall()
        proveedores = []

    movimientos = db.execute('''
        SELECT m.*, art.nombre as nombre_articulo, u.nombre as nombre_usuario
        FROM movimientos m
        JOIN articulos art ON m.id_articulo = art.id_articulo
        LEFT JOIN usuarios u ON m.id_usuario = u.id_usuario
        WHERE art.nombre LIKE ? OR art.num_parte LIKE ?
        ORDER BY m.fecha_hora DESC LIMIT 10
    ''', (f'%{q}%', f'%{q}%')).fetchall()

    db.close()
    return render_template('buscar.html', q=q, articulos=articulos,
                           proveedores=proveedores, movimientos=movimientos)


# ─── ESCANEO QR DIRECTO ───────────────────────────────────────────────────────

@app.route('/scan/<codigo>')
def scan_qr(codigo):
    db = get_db()
    # Buscar por codigo_qr o por num_parte
    articulo = db.execute('''
        SELECT a.*, p.nombre as nombre_proveedor 
        FROM articulos a 
        JOIN proveedores p ON a.id_proveedor = p.id_proveedor 
        WHERE (a.codigo_qr = %s OR a.num_parte = %s) AND a.activo = 1
    ''', (codigo, codigo)).fetchone()
    db.close()
    if not articulo:
        return render_template('scan_notfound.html', codigo=codigo)
    return render_template('scan_formulario.html', articulo=articulo)

@app.route('/scan/registrar', methods=['POST'])
def scan_registrar():
    data = request.get_json()
    codigo_qr = data.get('codigo_qr')
    tipo = data.get('tipo', 'SALIDA')
    cantidad = int(data.get('cantidad', 1))
    observaciones = data.get('observaciones', '')

    db = get_db()
    articulo = db.execute(
        'SELECT * FROM articulos WHERE codigo_qr = %s AND activo = 1', (codigo_qr,)
    ).fetchone()

    if not articulo:
        db.close()
        return jsonify({'status': 'error', 'message': 'Artículo no encontrado.'})

    # Solo admin puede hacer entradas desde escaneo directo
    rol = session.get('rol', 'kiosco')
    if tipo == 'ENTRADA' and rol not in ['admin']:
        db.close()
        return jsonify({'status': 'error', 'message': 'Solo el administrador puede registrar entradas.'})

    if tipo == 'SALIDA' and articulo['stock_actual'] < cantidad:
        db.close()
        return jsonify({'status': 'error', 'message': 'Stock insuficiente.'})

    db.execute('''
        INSERT INTO movimientos (id_articulo, tipo, cantidad, observaciones, modo_kiosco)
        VALUES (%s, %s, %s, %s, 1)
    ''', (articulo['id_articulo'], tipo, cantidad, observaciones))

    nuevo_stock = articulo['stock_actual'] + cantidad if tipo == 'ENTRADA' else articulo['stock_actual'] - cantidad
    db.execute('UPDATE articulos SET stock_actual = %s WHERE id_articulo = %s',
               (nuevo_stock, articulo['id_articulo']))

    if nuevo_stock <= articulo['stock_minimo']:
        db.execute('INSERT INTO alertas (id_articulo, tipo_alerta) VALUES (%s, %s)',
                   (articulo['id_articulo'], 'STOCK_MINIMO'))

    db.commit()
    db.close()
    return jsonify({
        'status': 'success',
        'message': f'Registrado correctamente. Stock actual: {nuevo_stock}',
        'stock_actual': nuevo_stock
    })

@app.route('/articulos/actualizar_qr')
@login_requerido
@solo_admin
def actualizar_qr():
    base_url = request.host_url.rstrip('/')
    db = get_db()
    articulos = db.execute('SELECT * FROM articulos WHERE activo=1').fetchall()
    for art in articulos:
        nuevo_qr = f"{base_url}/scan/{art['num_parte']}"
        db.execute('UPDATE articulos SET codigo_qr = %s WHERE id_articulo = %s',
                   (nuevo_qr, art['id_articulo']))
    db.commit()
    db.close()
    flash(f'QR actualizados para {len(articulos)} artículos.', 'success')
    return redirect(url_for('articulos'))

# ─── ÁREAS, PUESTOS Y MOTIVOS ─────────────────────────────────────────────────

@app.route('/configuracion/listas', methods=['GET', 'POST'])
@login_requerido
@solo_admin
def configuracion_listas():
    db = get_db()
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        nombre = request.form.get('nombre', '').strip()
        accion = request.form.get('accion')
        id_item = request.form.get('id_item')

        if accion == 'agregar' and nombre:
            if tipo == 'area':
                db.execute('INSERT INTO areas (nombre) VALUES (%s)', (nombre,))
            elif tipo == 'puesto':
                db.execute('INSERT INTO puestos (nombre) VALUES (%s)', (nombre,))
            elif tipo == 'motivo':
                db.execute('INSERT INTO motivos (nombre) VALUES (%s)', (nombre,))
            flash(f'Agregado correctamente.', 'success')

        elif accion == 'desactivar' and id_item:
            if tipo == 'area':
                db.execute('UPDATE areas SET activo=0 WHERE id_area=%s', (id_item,))
            elif tipo == 'puesto':
                db.execute('UPDATE puestos SET activo=0 WHERE id_puesto=%s', (id_item,))
            elif tipo == 'motivo':
                db.execute('UPDATE motivos SET activo=0 WHERE id_motivo=%s', (id_item,))
            flash('Eliminado correctamente.', 'success')

        db.commit()
        db.close()
        return redirect(url_for('configuracion_listas'))

    areas = db.execute('SELECT * FROM areas WHERE activo=1 ORDER BY nombre').fetchall()
    puestos = db.execute('SELECT * FROM puestos WHERE activo=1 ORDER BY nombre').fetchall()
    motivos = db.execute('SELECT * FROM motivos WHERE activo=1 ORDER BY nombre').fetchall()
    db.close()
    return render_template('configuracion_listas.html', areas=areas, puestos=puestos, motivos=motivos)

@app.route('/api/listas')
def api_listas():
    db = get_db()
    areas = [r['nombre'] for r in db.execute('SELECT nombre FROM areas WHERE activo=1 ORDER BY nombre').fetchall()]
    puestos = [r['nombre'] for r in db.execute('SELECT nombre FROM puestos WHERE activo=1 ORDER BY nombre').fetchall()]
    motivos = [r['nombre'] for r in db.execute('SELECT nombre FROM motivos WHERE activo=1 ORDER BY nombre').fetchall()]
    db.close()
    return jsonify({'areas': areas, 'puestos': puestos, 'motivos': motivos})
# ─── INICIO ───────────────────────────────────────────────────────────────────

@app.route('/api/articulo/<codigo>')
def api_articulo(codigo):
    db = get_db()
    art = db.execute(
        'SELECT * FROM articulos WHERE codigo_qr = ? AND activo = 1', (codigo,)
    ).fetchone()
    db.close()
    if art:
        return jsonify({'found': True, 'nombre': art['nombre'],
                        'stock_actual': art['stock_actual'], 'unidad_medida': art['unidad_medida']})
    return jsonify({'found': False})

if __name__ == '__main__':
    init_db()
    init_pool()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)